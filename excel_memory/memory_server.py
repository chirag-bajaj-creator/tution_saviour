from mcp.server.fastmcp import FastMCP
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

mcp = FastMCP("excel-memory")

XLSX_FILE = "memory.xlsx"
SHEET_NAME = "Memory"


def ensure_workbook():
    if not os.path.exists(XLSX_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["date", "prompt", "claude_response", "memory_written"])
        wb.save(XLSX_FILE)
    else:
        wb = load_workbook(XLSX_FILE)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(["date", "prompt", "claude_response", "memory_written"])
            wb.save(XLSX_FILE)


ensure_workbook()


@mcp.tool()
def add_memory(prompt: str, claude_response: str, memory_written: str = "") -> str:
    """Append a new memory row to the Excel file."""
    wb = load_workbook(XLSX_FILE)
    ws = wb[SHEET_NAME]

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        prompt,
        claude_response,
        memory_written
    ])

    wb.save(XLSX_FILE)
    return "Memory saved to memory.xlsx"


@mcp.tool()
def get_recent_memory(limit: int = 5) -> str:
    """Return the latest memory rows."""
    wb = load_workbook(XLSX_FILE, read_only=True)
    ws = wb[SHEET_NAME]

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return "No memory found."

    recent = rows[-limit:]
    output = []

    for row in recent:
        date, prompt, response, memory = row
        output.append(f"[{date}] Prompt: {prompt} | Memory: {memory}")

    return "\n".join(output)


@mcp.tool()
def search_memory(keyword: str, limit: int = 10) -> str:
    """Search memory rows by keyword."""
    wb = load_workbook(XLSX_FILE, read_only=True)
    ws = wb[SHEET_NAME]

    matches = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date, prompt, response, memory = row
        combined = f"{prompt or ''} {response or ''} {memory or ''}".lower()

        if keyword.lower() in combined:
            matches.append(f"[{date}] Prompt: {prompt} | Memory: {memory}")

        if len(matches) >= limit:
            break

    if not matches:
        return "No matching memory found."

    return "\n".join(matches)


if __name__ == "__main__":
    # For local Claude/CLI-style hosts, stdio is the normal transport
    mcp.run()